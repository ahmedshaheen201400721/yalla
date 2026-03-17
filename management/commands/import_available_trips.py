# -*- coding: utf-8 -*-
"""
Management command to import AvailableTrip records from an Excel file.

Usage:
    uv run python manage.py import_available_trips <path_to_xlsx>
"""
import os
import openpyxl
from django.core.management.base import BaseCommand, CommandError

from modules.base.services.import_export_service import ImportExportService


FIELD_MAPPING = {
    'S': '',                              # skip — serial number
    'Name': 'name',
    'Supplier': 'supplier',               # FK — Partners pre-created in pre-processing step
    'Destination': 'destination',
    'Short Description': 'note',          # maps to note TextField
    'Sell PRC Ad.': 'sell_prc_adult',
    'Sell PRC Ch.': 'sell_prc_child',
    'Net PRC Ad.': 'net_prc_adult',
    'Net PRC Ch.': 'net_prc_child',
    'Min. Markup': 'min_markup',
    'Sea Sickness': 'motion_sickness',
    'Weather Sensitivity': 'weather_sensitivity',
    'Children Eligibility': 'children_eligibility',
    'Action Adrenaline': 'action_adrenaline',
    'Romantic/Honeymoon': 'romantic_honeymoon',
    'Smoker Friendly': 'smoker_friendly',
    'Price Range': 'price_range',
    'Lunch': 'lunch',
    'National Park': 'national_park',
    'Stop/Act.': 'stop_activity',
    'Duration': 'duration',
    'Time': 'time',
    'Act. Type': 'activity_type',
    'Quality': 'quality',
    'Serivce Onboard': 'service_onboard',  # typo in Excel preserved
    'Stability': 'stability',
    'Mobility': 'mobility',
    'Boat View': 'boat_view',
    'Water Act.': 'water_activity',
    'No. of PAX': 'no_of_pax',
    'Longtail Boat': 'longtail_boat',
    'WhatsApp Catalog Link': 'whatsapp_catalog_link',
    'Video': 'video_link',
    'Album': 'album',
    'Note': 'note',
}

SUPPLIER_COLUMN = 'Supplier'
SHEET_NAME = 'Trips'


class Command(BaseCommand):
    help = 'Import AvailableTrip records from an Excel (.xlsx) file'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', type=str, help='Path to the .xlsx file')

    def handle(self, *args, **options):
        xlsx_path = options['xlsx_path']

        if not os.path.exists(xlsx_path):
            raise CommandError(f"File not found: {xlsx_path}")

        # Step 1 — Pre-create missing suppliers
        self.stdout.write('Pre-processing suppliers...')
        self._ensure_suppliers(xlsx_path)

        # Step 2 — Run import
        self.stdout.write('Starting import...')
        service = ImportExportService('AvailableTrip', 'yalla_thailand')

        # Override sheet name since our file uses 'Trips' not the default 'Data'
        from modules.base.services.format_handlers.excel_handler import ExcelHandler
        service.format_handlers['xlsx'] = ExcelHandler(sheet_name=SHEET_NAME)

        with open(xlsx_path, 'rb') as f:
            file_content = f.read()

        result = service.import_data(
            file_content=file_content,
            file_format='xlsx',
            field_mapping=FIELD_MAPPING,
            skip_invalid_rows=True,
        )

        # Step 3 — Report (result is an ImportExportJob instance)
        job = result
        self.stdout.write(self.style.SUCCESS(
            f"Import complete — status: {job.status} | "
            f"total: {job.total_records} | success: {job.success_records} | errors: {job.error_records}"
        ))

        if job.error_records:
            self.stdout.write(self.style.WARNING(f"{job.error_records} row(s) had errors (see job id={job.pk} for details)"))

    def _ensure_suppliers(self, xlsx_path):
        """Read all unique supplier names from the Excel and get_or_create Partners."""
        from modules.base.models.partner import Partner

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(
                f"Sheet '{SHEET_NAME}' not found. Available sheets: {wb.sheetnames}"
            )

        ws = wb[SHEET_NAME]
        rows = iter(ws.rows)

        # Find header row
        header_row = next(rows, None)
        if header_row is None:
            raise CommandError("Excel file appears to be empty.")

        headers = [cell.value for cell in header_row]
        try:
            supplier_col_idx = headers.index(SUPPLIER_COLUMN)
        except ValueError:
            self.stdout.write(self.style.WARNING(
                f"Column '{SUPPLIER_COLUMN}' not found in headers — skipping supplier pre-creation"
            ))
            wb.close()
            return

        supplier_names = set()
        for row in rows:
            val = row[supplier_col_idx].value
            if val and str(val).strip():
                supplier_names.add(str(val).strip())

        wb.close()

        created_count = 0
        for name in sorted(supplier_names):
            _, created = Partner.objects.get_or_create(
                name=name,
                defaults={'is_company': True},
            )
            if created:
                created_count += 1

        self.stdout.write(
            f"  Suppliers: {len(supplier_names)} unique names found, "
            f"{created_count} new Partner records created."
        )
